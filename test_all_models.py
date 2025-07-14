#!/usr/bin/env python3
"""
测试所有best.pt或last.pt模型并保存详细结果
包括Precision, Recall, mAP, 预测错误的图片等
支持通过命令行参数选择模型类型.
"""

import argparse
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# 添加项目根目录到路径
FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))


def get_available_train_folders():
    """获取所有可用的训练文件夹."""
    runs_dir = Path("runs")
    train_folders = []

    if runs_dir.exists():
        for folder in runs_dir.iterdir():
            if (
                folder.is_dir()
                and ("train" in folder.name or "rail_train" in folder.name)
                and "epoch" in folder.name
                and "_test_" not in folder.name
            ):  # 排除测试结果文件夹
                # 检查是否包含模型文件
                has_models = False
                for exp_dir in folder.iterdir():
                    if exp_dir.is_dir():
                        weights_dir = exp_dir / "weights"
                        if (
                            weights_dir.exists()
                            and (weights_dir / "best.pt").exists()
                            or (weights_dir / "last.pt").exists()
                        ):
                            has_models = True
                            break
                if has_models:
                    train_folders.append(folder.name)

    return sorted(train_folders)


def get_all_models(model_type="best", train_folder="train200epoch"):
    """获取所有指定类型的模型路径.

    Args:
        model_type (str): 模型类型，'best' 或 'last'
        train_folder (str): 训练文件夹名称，如 'train200epoch', 'train300epoch'

    Returns:
        list: 包含模型信息的列表
    """
    train_dir = Path(f"runs/{train_folder}")
    models = []

    if not train_dir.exists():
        print(f"警告: 训练目录不存在 {train_dir}")
        return models

    model_filename = f"{model_type}.pt"

    for exp_dir in train_dir.iterdir():
        if exp_dir.is_dir():
            weights_dir = exp_dir / "weights"
            model_pt = weights_dir / model_filename
            if model_pt.exists():
                models.append(
                    {"name": exp_dir.name, "path": str(model_pt), "type": model_type, "train_folder": train_folder}
                )

    return models


def parse_val_output(output_text):
    """从val.py的输出中解析性能指标."""
    results = {}

    # 查找包含性能指标的行
    lines = output_text.split("\n")

    # 调试：保存原始输出用于分析
    results["_debug_output"] = output_text

    for line in lines:
        # 查找包含 "all" 和性能指标的行，去除ANSI颜色代码
        clean_line = re.sub(r"\x1b\[[0-9;]*m", "", line)  # 移除ANSI颜色代码

        # 查找包含 "all" 的行，格式应该是: "all 779 1648 0.87 0.851 0.887 0.536"
        if clean_line.strip().startswith("all"):
            # 分割行并提取数值
            parts = clean_line.strip().split()
            if len(parts) >= 7:  # all + 6个数值
                try:
                    results["precision"] = float(parts[3])
                    results["recall"] = float(parts[4])
                    results["map50"] = float(parts[5])
                    results["map50_95"] = float(parts[6])
                except (ValueError, IndexError):
                    continue

        # 查找NO-Safety Vest类别的指标
        # 实际格式: "        NO-Safety Vest        779        361      0.846      0.798      0.834      0.403"
        elif "NO-Safety Vest" in clean_line:
            parts = clean_line.strip().split()
            if len(parts) >= 7:  # NO-Safety + Vest + 图片数 + 实例数 + 4个指标
                try:
                    # 格式: "NO-Safety Vest 779 361 0.846 0.798 0.834 0.403"
                    # 索引:     0       1   2   3     4     5     6     7
                    results["no_safety_vest_precision"] = float(parts[4])
                    results["no_safety_vest_recall"] = float(parts[5])
                    results["no_safety_vest_map50"] = float(parts[6])
                    results["no_safety_vest_map50_95"] = float(parts[7])
                    results["_debug_no_vest_line"] = clean_line
                    results["_debug_no_vest_parts"] = parts
                except (ValueError, IndexError):
                    continue

    return results


def calculate_ra_map(map50_95, no_safety_vest_recall):
    """
    计算RA-mAP指标.

    RA-mAP = 0.4 * mAP@0.5:0.95 + 0.6 * NO-Safety Vest Recall

    Args:
        map50_95 (float): mAP@0.5:0.95 值
        no_safety_vest_recall (float): NO-Safety Vest 召回率

    Returns:
        float: RA-mAP 值，如果输入无效则返回None
    """
    if isinstance(map50_95, (int, float)) and isinstance(no_safety_vest_recall, (int, float)):
        return 0.4 * map50_95 + 0.6 * no_safety_vest_recall
    return None


def run_validation(model_info, output_dir, data_yaml, conf_thres=0.001, iou_thres=0.6, eval_split="val"):
    """运行验证并保存结果.

    Args:
        model_info (dict): 模型信息字典
        output_dir (Path): 输出目录
        data_yaml (str): 数据集配置文件路径
        conf_thres (float): 置信度阈值
        iou_thres (float): IoU阈值
        eval_split (str): 评估数据集选择，'val' 或 'test'
    """
    model_name = model_info["name"]
    model_path = model_info["path"]

    print(f"正在测试模型: {model_name}")

    # 运行验证命令 - 根据eval_split参数选择数据集
    cmd = [
        "python",
        "val.py",
        "--weights",
        model_path,
        "--data",
        data_yaml,
        "--img",
        "640",
        "--batch",
        "16",
        "--conf",
        str(conf_thres),
        "--iou",
        str(iou_thres),
        "--task",
        eval_split,  # 使用指定的数据集 (val 或 test)
        "--save-txt",
        "--save-conf",
        "--save-json",
        "--project",
        str(output_dir),
        "--name",
        model_name,
        "--exist-ok",
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        print(f"模型 {model_name} 验证完成")

        # 合并stdout和stderr的输出
        full_output = result.stdout + result.stderr

        # 解析输出中的性能指标
        metrics = parse_val_output(full_output)
        return True, full_output, metrics
    except subprocess.CalledProcessError as e:
        print(f"模型 {model_name} 验证失败: {e}")
        return False, e.stderr, {}


def parse_results(model_name, output_dir):
    """解析验证结果."""
    results_dir = Path(output_dir) / model_name

    # 读取results.csv
    results_csv = results_dir / "results.csv"
    if results_csv.exists():
        df = pd.read_csv(results_csv)
        # 获取最后一行的结果（最新的验证结果）
        latest_results = df.iloc[-1].to_dict()
        return latest_results
    else:
        print(f"警告: 未找到结果文件 {results_csv}")
        return None


def save_error_images(model_name, output_dir, base_output_dir, data_yaml, eval_split="test"):
    """保存预测错误的图片."""
    results_dir = Path(output_dir) / model_name
    error_dir = Path(base_output_dir) / "error_images" / model_name
    error_dir.mkdir(parents=True, exist_ok=True)

    # 从data.yaml文件中获取对应数据集路径
    import yaml

    try:
        with open(data_yaml) as f:
            data_config = yaml.safe_load(f)

        # 根据eval_split参数获取对应的数据集路径
        if eval_split == "test":
            dataset_path_key = "test"
        else:
            dataset_path_key = "val"

        val_path = data_config.get(dataset_path_key, data_config.get("valid", ""))
        if val_path:
            # 检查路径是否为绝对路径
            val_path_obj = Path(val_path)
            if val_path_obj.is_absolute():
                dataset_path = val_path_obj
            else:
                # 如果是相对路径，检查是否以项目根目录为基准
                if val_path.startswith("data/"):
                    # 直接使用该路径（相对于项目根目录）
                    dataset_path = Path(val_path)
                else:
                    # 相对于data.yaml文件的目录
                    data_yaml_dir = Path(data_yaml).parent
                    dataset_path = data_yaml_dir / val_path
        else:
            print(f"警告: 无法从 {data_yaml} 中获取验证集路径")
            return
    except Exception as e:
        print(f"警告: 读取数据配置文件失败 {data_yaml}: {e}")
        return

    # 验证集路径应该指向包含images和labels的目录
    # 如果dataset_path指向images目录，需要获取其父目录
    if dataset_path.name == "images":
        dataset_base = dataset_path.parent
        labels_path = dataset_base / "labels"
        images_path = dataset_path
    else:
        # 如果dataset_path是基础目录，添加子目录
        labels_path = dataset_path / "labels"
        images_path = dataset_path / "images"

    # 获取所有预测结果文件
    pred_labels_dir = results_dir / "labels"
    if not pred_labels_dir.exists():
        print(f"警告: 预测标签目录不存在 {pred_labels_dir}")
        return

    pred_files = list(pred_labels_dir.glob("*.txt"))
    error_images = set()

    print(f"正在分析 {len(pred_files)} 个预测结果文件...")
    print(f"真实标签路径: {labels_path}")
    print(f"图片路径: {images_path}")

    for pred_file in pred_files:
        # 获取对应的真实标签文件
        gt_file = labels_path / pred_file.name

        if not gt_file.exists():
            print(f"警告: 未找到对应的真实标签文件 {gt_file}")
            continue

        # 读取预测结果
        pred_boxes = []
        try:
            with open(pred_file) as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) >= 5:
                        class_id = int(parts[0])
                        x_center = float(parts[1])
                        y_center = float(parts[2])
                        width = float(parts[3])
                        height = float(parts[4])
                        conf = float(parts[5]) if len(parts) > 5 else 1.0
                        pred_boxes.append([class_id, x_center, y_center, width, height, conf])
        except Exception as e:
            print(f"读取预测文件失败 {pred_file}: {e}")
            continue

        # 读取真实标签
        gt_boxes = []
        try:
            with open(gt_file) as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) >= 5:
                        class_id = int(parts[0])
                        x_center = float(parts[1])
                        y_center = float(parts[2])
                        width = float(parts[3])
                        height = float(parts[4])
                        gt_boxes.append([class_id, x_center, y_center, width, height])
        except Exception as e:
            print(f"读取真实标签文件失败 {gt_file}: {e}")
            continue

        # 检查是否有预测错误
        has_error = False

        # 检查假阳性（预测了但实际没有的框）
        if len(pred_boxes) > len(gt_boxes):
            has_error = True

        # 检查假阴性（实际有但预测没有的框）
        elif len(pred_boxes) < len(gt_boxes):
            has_error = True

        # 检查类别错误或位置错误
        else:
            # 简单的IoU检查（这里使用简化的检查方法）
            for pred_box in pred_boxes:
                pred_class = pred_box[0]
                pred_center_x, pred_center_y = pred_box[1], pred_box[2]
                pred_w, pred_h = pred_box[3], pred_box[4]

                # 检查是否有匹配的真实框
                matched = False
                for gt_box in gt_boxes:
                    gt_class = gt_box[0]
                    gt_center_x, gt_center_y = gt_box[1], gt_box[2]
                    gt_w, gt_h = gt_box[3], gt_box[4]

                    # 检查类别是否匹配
                    if pred_class == gt_class:
                        # 检查位置是否接近（简化的IoU检查）
                        center_dist = ((pred_center_x - gt_center_x) ** 2 + (pred_center_y - gt_center_y) ** 2) ** 0.5
                        size_diff = abs(pred_w * pred_h - gt_w * gt_h)

                        # 如果中心点距离小于阈值且大小差异不大，认为是匹配的
                        if center_dist < 0.1 and size_diff < 0.1:
                            matched = True
                            break

                if not matched:
                    has_error = True
                    break

        # 如果有错误，将对应的图片添加到错误图片列表
        if has_error:
            # 从文件名推断原始图片名
            # 预测文件名格式: original_name_jpg.rf.hash.txt
            # 需要找到对应的图片文件
            base_name = pred_file.stem  # 去掉.txt
            # 查找对应的图片文件
            image_extensions = [".jpg", ".jpeg", ".png"]
            for ext in image_extensions:
                # 尝试不同的图片文件名模式
                possible_names = [
                    base_name + ext,
                    base_name.replace("_jpg.rf.", ".") + ext,
                    base_name.split("_jpg.rf.")[0] + ext,
                ]

                for possible_name in possible_names:
                    # 检查在数据集目录中是否有对应的图片
                    dataset_image = images_path / possible_name
                    if dataset_image.exists():
                        error_images.add(dataset_image)
                        break

    # 复制错误图片到error_images目录
    copied_count = 0
    for img_path in error_images:
        try:
            # 复制图片
            shutil.copy2(img_path, error_dir / img_path.name)
            copied_count += 1
        except Exception as e:
            print(f"复制图片失败 {img_path}: {e}")

    print(f"已保存 {copied_count} 张预测错误的图片到 {error_dir}")

    # 如果没有找到错误图片，创建一个说明文件
    if copied_count == 0:
        with open(error_dir / "no_errors_found.txt", "w", encoding="utf-8") as f:
            f.write("未发现预测错误的图片\n")
            f.write("所有预测结果都是正确的\n")
        print("未发现预测错误，已创建说明文件")


def create_summary_report(
    models_results,
    output_dir,
    model_type="best",
    train_folder="train200epoch",
    data_yaml="data/SafetyVests.v6/data.yaml",
    eval_split="val",
):
    """创建汇总报告.

    Args:
        models_results (dict): 模型结果字典
        output_dir (Path): 输出目录
        model_type (str): 模型类型，'best' 或 'last'
        train_folder (str): 训练文件夹名称
        data_yaml (str): 数据集配置文件路径
        eval_split (str): 评估数据集，'val' 或 'test'
    """
    summary_file = Path(output_dir) / "summary_report.txt"

    with open(summary_file, "w", encoding="utf-8") as f:
        f.write("=" * 100 + "\n")
        f.write(f"YOLOv5 {model_type.upper()}模型测试汇总报告\n")
        f.write("=" * 100 + "\n")
        f.write(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"测试数据集: {data_yaml}\n")
        f.write(f"评估数据集: {eval_split.upper()} ({'验证集' if eval_split == 'val' else '测试集'})\n")
        f.write(f"训练文件夹: runs/{train_folder}\n")
        f.write(f"模型类型: {model_type}.pt\n")
        f.write(f"模型数量: {len(models_results)}\n\n")

        # 创建整体性能结果表格
        f.write("整体模型性能对比:\n")
        f.write("-" * 120 + "\n")
        f.write(
            f"{'模型名称':<20} {'Precision':<12} {'Recall':<12} {'mAP@0.5':<12} {'mAP@0.5:0.95':<15} {'RA-mAP':<12}\n"
        )
        f.write("-" * 120 + "\n")

        for model_name, results in models_results.items():
            if results:
                precision = results.get("precision", "N/A")
                recall = results.get("recall", "N/A")
                map50 = results.get("map50", "N/A")
                map50_95 = results.get("map50_95", "N/A")
                no_vest_recall = results.get("no_safety_vest_recall", "N/A")

                # 计算RA-mAP
                ra_map = calculate_ra_map(map50_95, no_vest_recall)

                if isinstance(precision, (int, float)):
                    ra_map_str = f"{ra_map:.4f}" if ra_map is not None else "N/A"
                    f.write(
                        f"{model_name:<20} {precision:<12.4f} {recall:<12.4f} {map50:<12.4f} {map50_95:<15.4f} {ra_map_str:<12}\n"
                    )
                else:
                    f.write(f"{model_name:<20} {'N/A':<12} {'N/A':<12} {'N/A':<12} {'N/A':<15} {'N/A':<12}\n")
            else:
                f.write(f"{model_name:<20} {'N/A':<12} {'N/A':<12} {'N/A':<12} {'N/A':<15} {'N/A':<12}\n")

        f.write("-" * 120 + "\n\n")

        # 创建NO-Safety Vest类别专门的性能表格
        f.write("NO-Safety Vest 类别性能对比:\n")
        f.write("-" * 100 + "\n")
        f.write(f"{'模型名称':<20} {'Precision':<12} {'Recall':<12} {'mAP@0.5':<12} {'mAP@0.5:0.95':<15}\n")
        f.write("-" * 100 + "\n")

        for model_name, results in models_results.items():
            if results:
                no_vest_precision = results.get("no_safety_vest_precision", "N/A")
                no_vest_recall = results.get("no_safety_vest_recall", "N/A")
                no_vest_map50 = results.get("no_safety_vest_map50", "N/A")
                no_vest_map50_95 = results.get("no_safety_vest_map50_95", "N/A")

                if isinstance(no_vest_precision, (int, float)):
                    f.write(
                        f"{model_name:<20} {no_vest_precision:<12.4f} {no_vest_recall:<12.4f} {no_vest_map50:<12.4f} {no_vest_map50_95:<15.4f}\n"
                    )
                else:
                    f.write(f"{model_name:<20} {'N/A':<12} {'N/A':<12} {'N/A':<12} {'N/A':<15}\n")
            else:
                f.write(f"{model_name:<20} {'N/A':<12} {'N/A':<12} {'N/A':<12} {'N/A':<15}\n")

        f.write("-" * 100 + "\n\n")

        # 找出最佳模型（基于整体mAP@0.5）
        best_model_overall = None
        best_map50_overall = 0

        for model_name, results in models_results.items():
            if results:
                map50 = results.get("map50", 0)
                if isinstance(map50, (int, float)) and map50 > best_map50_overall:
                    best_map50_overall = map50
                    best_model_overall = model_name

        # 找出NO-Safety Vest召回率最佳的模型
        best_model_no_vest_recall = None
        best_no_vest_recall = 0

        for model_name, results in models_results.items():
            if results:
                no_vest_recall = results.get("no_safety_vest_recall", 0)
                if isinstance(no_vest_recall, (int, float)) and no_vest_recall > best_no_vest_recall:
                    best_no_vest_recall = no_vest_recall
                    best_model_no_vest_recall = model_name

        # 找出RA-mAP最佳的模型
        best_model_ra_map = None
        best_ra_map = 0

        for model_name, results in models_results.items():
            if results:
                map50_95 = results.get("map50_95", 0)
                no_vest_recall = results.get("no_safety_vest_recall", 0)
                ra_map = calculate_ra_map(map50_95, no_vest_recall)
                if ra_map is not None and ra_map > best_ra_map:
                    best_ra_map = ra_map
                    best_model_ra_map = model_name

        f.write("最佳模型分析:\n")
        f.write("-" * 50 + "\n")
        if best_model_overall:
            f.write(f"整体最佳模型 (基于mAP@0.5): {best_model_overall}\n")
            f.write(f"最佳整体mAP@0.5: {best_map50_overall:.4f}\n\n")

        if best_model_no_vest_recall:
            f.write(f"NO-Safety Vest召回率最佳模型: {best_model_no_vest_recall}\n")
            f.write(f"最佳NO-Safety Vest召回率: {best_no_vest_recall:.4f}\n\n")

        if best_model_ra_map:
            f.write(f"RA-mAP最佳模型: {best_model_ra_map}\n")
            f.write(f"最佳RA-mAP值: {best_ra_map:.4f}\n")
            f.write("RA-mAP计算公式: 0.4 × mAP@0.5:0.95 + 0.6 × NO-Safety Vest Recall\n\n")

        f.write("详细结果文件位置:\n")
        for model_name in models_results.keys():
            f.write(f"- {model_name}: {output_dir}/{model_name}/\n")

        f.write(f"\n错误图片位置: {output_dir}/error_images/\n")

    print(f"汇总报告已保存到: {summary_file}")


def create_performance_table(
    models_results, output_dir, model_type="best", train_folder="train200epoch", eval_split="test"
):
    """
    创建性能对比表格 (CSV和Excel格式).

    Args:
        models_results (dict): 模型结果字典
        output_dir (Path): 输出目录
        model_type (str): 模型类型
        train_folder (str): 训练文件夹名称
        eval_split (str): 评估数据集类型
    """
    # 准备数据
    table_data = []

    for model_name, results in models_results.items():
        if results:
            # 获取基础指标
            map50 = results.get("map50", None)
            map50_95 = results.get("map50_95", None)
            no_vest_recall = results.get("no_safety_vest_recall", None)

            # 计算RA-mAP
            ra_map = calculate_ra_map(map50_95, no_vest_recall)

            # 格式化数据
            row = {
                "模型名称": model_name,
                "mAP@0.5": f"{map50:.4f}" if isinstance(map50, (int, float)) else "N/A",
                "mAP@0.5:0.95": f"{map50_95:.4f}" if isinstance(map50_95, (int, float)) else "N/A",
                "NO-Safety Vest Recall": f"{no_vest_recall:.4f}" if isinstance(no_vest_recall, (int, float)) else "N/A",
                "RA-mAP": f"{ra_map:.4f}" if ra_map is not None else "N/A",
                # 添加原始数值用于排序
                "_map50_raw": map50 if isinstance(map50, (int, float)) else 0,
                "_map50_95_raw": map50_95 if isinstance(map50_95, (int, float)) else 0,
                "_no_vest_recall_raw": no_vest_recall if isinstance(no_vest_recall, (int, float)) else 0,
                "_ra_map_raw": ra_map if ra_map is not None else 0,
            }
        else:
            row = {
                "模型名称": model_name,
                "mAP@0.5": "N/A",
                "mAP@0.5:0.95": "N/A",
                "NO-Safety Vest Recall": "N/A",
                "RA-mAP": "N/A",
                "_map50_raw": 0,
                "_map50_95_raw": 0,
                "_no_vest_recall_raw": 0,
                "_ra_map_raw": 0,
            }

        table_data.append(row)

    # 按RA-mAP降序排序
    table_data.sort(key=lambda x: x["_ra_map_raw"], reverse=True)

    # 创建DataFrame
    df = pd.DataFrame(table_data)

    # 移除用于排序的原始数值列
    display_columns = ["模型名称", "mAP@0.5", "mAP@0.5:0.95", "NO-Safety Vest Recall", "RA-mAP"]
    df_display = df[display_columns].copy()

    # 保存CSV文件
    csv_file = output_dir / f"performance_comparison_{model_type}_{train_folder}_{eval_split}.csv"
    df_display.to_csv(csv_file, index=False, encoding="utf-8-sig")
    print(f"性能对比CSV表格已保存到: {csv_file}")

    # 保存Excel文件
    excel_file = output_dir / f"performance_comparison_{model_type}_{train_folder}_{eval_split}.xlsx"

    try:
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # 写入主表
            df_display.to_excel(writer, sheet_name="性能对比", index=False)

            # 获取工作表
            worksheet = writer.sheets["性能对比"]

            # 设置列宽
            column_widths = {
                "A": 25,  # 模型名称
                "B": 15,  # mAP@0.5
                "C": 18,  # mAP@0.5:0.95
                "D": 22,  # NO-Safety Vest Recall
                "E": 15,  # RA-mAP
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col in range(1, len(display_columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 设置数据行样式
            data_alignment = Alignment(horizontal="center", vertical="center")
            for row in range(2, len(df_display) + 2):
                for col in range(1, len(display_columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment

                    # 为最佳RA-mAP值添加高亮
                    if col == 5 and row == 2:  # RA-mAP列的第一行（最高值）
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            # 添加说明工作表
            info_data = [
                ["指标说明", ""],
                ["mAP@0.5", "在IoU阈值0.5下的平均精度"],
                ["mAP@0.5:0.95", "在IoU阈值0.5到0.95下的平均精度"],
                ["NO-Safety Vest Recall", "NO-Safety Vest类别的召回率"],
                ["RA-mAP", "新提出的综合指标"],
                ["", ""],
                ["RA-mAP计算公式", ""],
                ["RA-mAP = 0.4 × mAP@0.5:0.95 + 0.6 × NO-Safety Vest Recall", ""],
                ["", ""],
                ["说明", ""],
                ["- RA-mAP综合考虑了整体检测精度和NO-Safety Vest类别的召回率", ""],
                ["- mAP@0.5:0.95权重为0.4，体现整体检测能力", ""],
                ["- NO-Safety Vest Recall权重为0.6，突出安全背心检测的重要性", ""],
                ["- 表格按RA-mAP降序排列，值越高表示性能越好", ""],
            ]

            info_df = pd.DataFrame(info_data, columns=["项目", "说明"])
            info_df.to_excel(writer, sheet_name="指标说明", index=False)

            # 设置说明工作表样式
            info_worksheet = writer.sheets["指标说明"]
            info_worksheet.column_dimensions["A"].width = 50
            info_worksheet.column_dimensions["B"].width = 60

        print(f"性能对比Excel表格已保存到: {excel_file}")

    except Exception as e:
        print(f"保存Excel文件时出错: {e}")
        print("CSV文件已成功保存")


def parse_args():
    """解析命令行参数."""
    # 获取可用的训练文件夹
    available_folders = get_available_train_folders()

    parser = argparse.ArgumentParser(
        description="测试YOLOv5模型性能",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
可用的训练文件夹:
{chr(10).join(f"  - {folder}" for folder in available_folders) if available_folders else "  未找到训练文件夹"}

使用示例:
  python test_all_models.py --model-type best --train-folder train200epoch  # 默认使用测试集(论文指标)
  python test_all_models.py --model-type last --train-folder train300epoch --conf-thres 0.25
  python test_all_models.py --train-folder rail_train300epoch --data data/railroad-worker-detection/data.yaml
  python test_all_models.py --eval-split val --model-type best --train-folder train200epoch  # 开发阶段使用验证集
        """,
    )

    parser.add_argument(
        "--model-type",
        type=str,
        choices=["best", "last"],
        default="best",
        help="选择模型类型: best.pt 或 last.pt (默认: best)",
    )
    parser.add_argument(
        "--train-folder",
        type=str,
        default="train200epoch",
        help=f"训练文件夹名称 (默认: train200epoch)\n可用选项: {', '.join(available_folders) if available_folders else '无'}",
    )
    parser.add_argument(
        "--data",
        type=str,
        default="data/SafetyVests.v6/data.yaml",
        help="数据集配置文件路径 (默认: data/SafetyVests.v6/data.yaml)",
    )
    parser.add_argument("--conf-thres", type=float, default=0.001, help="置信度阈值 (默认: 0.001)")
    parser.add_argument("--iou-thres", type=float, default=0.6, help="IoU阈值 (默认: 0.6)")
    parser.add_argument("--list-folders", action="store_true", help="列出所有可用的训练文件夹并退出")
    parser.add_argument(
        "--eval-split",
        type=str,
        choices=["val", "test"],
        default="test",
        help="选择评估数据集: val (验证集) 或 test (测试集)。论文指标使用test (默认: test)",
    )

    args = parser.parse_args()

    # 如果用户要求列出文件夹，显示后退出
    if args.list_folders:
        print("可用的训练文件夹:")
        if available_folders:
            for folder in available_folders:
                print(f"  - {folder}")
        else:
            print("  未找到任何训练文件夹")
        sys.exit(0)

    # 验证训练文件夹是否存在
    if args.train_folder not in available_folders:
        print(f"错误: 训练文件夹 '{args.train_folder}' 不存在")
        print(f"可用的训练文件夹: {', '.join(available_folders) if available_folders else '无'}")
        sys.exit(1)

    # 验证数据集配置文件是否存在
    if not Path(args.data).exists():
        print(f"错误: 数据集配置文件 '{args.data}' 不存在")
        sys.exit(1)

    return args


def main():
    """主函数."""
    args = parse_args()
    model_type = args.model_type
    train_folder = args.train_folder

    print(f"开始测试 {train_folder} 文件夹下的所有{model_type}.pt模型...")

    # 获取所有模型
    models = get_all_models(model_type, train_folder)
    if not models:
        print(f"未找到任何{model_type}.pt模型文件在 runs/{train_folder} 目录下！")
        print("请检查目录是否存在以及是否包含模型文件。")
        return

    print(f"找到 {len(models)} 个{model_type}模型:")
    for model in models:
        print(f"  - {model['name']}: {model['path']}")

    # 创建输出目录 - 包含训练文件夹名称和评估数据集类型
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    eval_suffix = "test" if args.eval_split == "test" else "val"
    output_dir = Path(f"runs/{train_folder}_{eval_suffix}_{model_type}_{timestamp}")
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n训练文件夹: runs/{train_folder}")
    print(f"输出目录: {output_dir}")
    print(f"评估数据集: {args.eval_split.upper()} ({'验证集' if args.eval_split == 'val' else '测试集'})")
    print(f"置信度阈值: {args.conf_thres}")
    print(f"IoU阈值: {args.iou_thres}")

    # 存储所有模型的结果
    models_results = {}

    # 测试每个模型
    for i, model in enumerate(models, 1):
        print(f"\n[{i}/{len(models)}] 测试模型: {model['name']}")
        success, _, metrics = run_validation(
            model, output_dir, args.data, args.conf_thres, args.iou_thres, args.eval_split
        )

        if success:
            # 保存解析的指标
            models_results[model["name"]] = metrics

            # 保存错误图片
            save_error_images(model["name"], output_dir, output_dir, args.data, args.eval_split)

            print(f"模型 {model['name']} 测试完成")
            if metrics:
                print(
                    f"  整体性能指标: Precision={metrics.get('precision', 'N/A')}, "
                    f"Recall={metrics.get('recall', 'N/A')}, "
                    f"mAP@0.5={metrics.get('map50', 'N/A')}, "
                    f"mAP@0.5:0.95={metrics.get('map50_95', 'N/A')}"
                )

                # 显示NO-Safety Vest类别的召回率
                no_vest_recall = metrics.get("no_safety_vest_recall", "N/A")
                if no_vest_recall != "N/A":
                    print(f"  NO-Safety Vest召回率: {no_vest_recall:.4f}")
                    print(
                        f"  NO-Safety Vest其他指标: Precision={metrics.get('no_safety_vest_precision', 'N/A'):.4f}, "
                        f"mAP@0.5={metrics.get('no_safety_vest_map50', 'N/A'):.4f}, "
                        f"mAP@0.5:0.95={metrics.get('no_safety_vest_map50_95', 'N/A'):.4f}"
                    )

                    # 计算并显示RA-mAP
                    map50_95 = metrics.get("map50_95", "N/A")
                    if map50_95 != "N/A":
                        ra_map = calculate_ra_map(map50_95, no_vest_recall)
                        if ra_map is not None:
                            print(f"  RA-mAP (新指标): {ra_map:.4f}")
                            print(f"    计算公式: 0.4 × {map50_95:.4f} + 0.6 × {no_vest_recall:.4f} = {ra_map:.4f}")
                        else:
                            print("  RA-mAP: 无法计算")
                    else:
                        print("  RA-mAP: 无法计算 (缺少mAP@0.5:0.95)")
                else:
                    print("  NO-Safety Vest召回率: 未能解析")
                    print("  RA-mAP: 无法计算 (缺少NO-Safety Vest召回率)")
                    # 调试信息
                    if "_debug_no_vest_line" in metrics:
                        print(f"  调试: 找到NO-Safety Vest行: {metrics['_debug_no_vest_line']}")
                    elif "_debug_no_vest_line_fallback" in metrics:
                        print(f"  调试: 使用备选解析: {metrics['_debug_no_vest_line_fallback']}")
                    else:
                        print("  调试: 未找到NO-Safety Vest相关行，请检查输出格式")
        else:
            print(f"模型 {model['name']} 测试失败")
            models_results[model["name"]] = {}

    # 创建汇总报告
    create_summary_report(models_results, output_dir, model_type, train_folder, args.data, args.eval_split)

    # 创建性能对比表格
    create_performance_table(models_results, output_dir, model_type, train_folder, args.eval_split)

    # 保存详细结果到JSON
    json_file = output_dir / "detailed_results.json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(models_results, f, indent=2, ensure_ascii=False)

    print("\n测试完成！")
    print(f"所有结果已保存到: {output_dir}")
    print(f"汇总报告: {output_dir}/summary_report.txt")
    print(f"详细JSON结果: {json_file}")
    print(f"性能对比表格: {output_dir}/performance_comparison_{model_type}_{train_folder}_{args.eval_split}.csv")
    print(f"性能对比表格: {output_dir}/performance_comparison_{model_type}_{train_folder}_{args.eval_split}.xlsx")


if __name__ == "__main__":
    main()
